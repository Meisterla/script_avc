import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook


path = os.getcwd()
#df = pd.read_excel(r"C:\Users\Administrator\Desktop\数据分析\19Q4洗衣机线上-平台历史数据变更说明模板\季度变更文档往期.xlsx")
df = pd.read_excel(path + r"\季度变更文档往期.xlsx")

shuxing_list = df.columns.tolist()[6:]

for i in shuxing_list:
    df[i] = df[i].str.upper()

dictshuxing = {}
for i in shuxing_list:
    d = df[i].drop_duplicates(keep='first', inplace=False)
    dictshuxing[i] = len(d)

df_xiaoliang = df.pivot_table(index = '品类', columns = '周度', values = '销量', aggfunc = np.sum, margins=True)
df_xiaoliang2 = df_xiaoliang.iloc[0:-1,-9:]#元素一
df_xiaoe = df.pivot_table(index = '品类', columns = '周度', values = '销额', aggfunc = np.sum, margins=True)
df_xiaoe2 = df_xiaoe.iloc[0:-1,-9:]/10000#元素二
df_junjia = df_xiaoe/df_xiaoliang
df_junjia2 = df_junjia.iloc[0:-1,-9:]#元素三

sort_week = df_junjia2.columns[-2]

df_xiaoezhanbi = df.pivot_table(index = '品牌', columns = '周度', values = '销额', aggfunc = np.sum, margins=True).sort_values(sort_week,ascending=False).drop(labels = 'All')

for i in range(-9,0):
    df_xiaoezhanbi[df_xiaoezhanbi.columns[i]] = (df_xiaoezhanbi[df_xiaoezhanbi.columns[i]] / df_xiaoezhanbi[df_xiaoezhanbi.columns[i]].sum())

df_topxiaoezhanbi = df_xiaoezhanbi.iloc[:10,-9:]#元素四

df_xiaoexiaoliangzhanbi = df.pivot_table(index = '品牌'
                                      , columns = '周度'
                                      , values = ['销额','销量']
                                      , aggfunc = {'销额':np.sum,'销量':np.sum}
                                      , margins=True).sort_values(('销额',sort_week),ascending=False).drop(labels = 'All').sort_values(('销额',sort_week),ascending=False)

df_junjiazhanbi = df_xiaoexiaoliangzhanbi['销额']/df_xiaoexiaoliangzhanbi['销量']
df_topjunjiazhanbi = df_junjiazhanbi.iloc[:10,-9:]#元素五

wb = load_workbook(path + r"\平台历史数据变更说明模板.xlsx")
wb.get_sheet_names()

ws = wb.get_sheet_by_name("平台数据变更说明表")

weeklist = [*df_xiaoliang2.columns.values[:-1]]*4
indexlist = []
columns1 = ['C','D','E','F','G','H','I','J']
index1 = [20,32,66,100]
for i in index1:
    for j in columns1:
        indexlist.append(j+str(i))

for i in range(0,len(weeklist)):
    ws[indexlist[i]] = weeklist[i]

xiaolianglist = df_xiaoliang2.values.tolist()[0]
indexlist2 = ['C21','D21','E21','F21','G21','H21','I21','J21','K21']#销量
for i in range(0,len(xiaolianglist)):
    ws[indexlist2[i]] = xiaolianglist[i]

xiaoelist = df_xiaoe2.values.tolist()[0]
indexlist3 = ['C22','D22','E22','F22','G22','H22','I22','J22','K22']#销额
for i in range(0,len(xiaoelist)):
    ws[indexlist3[i]] = xiaoelist[i]

junjialist = df_junjia2.values.tolist()[0]
indexlist4 = ['C23','D23','E23','F23','G23','H23','I23','J23','K23']#均价
for i in range(0,len(junjialist)):
    ws[indexlist4[i]] = junjialist[i]

toplist = df_topxiaoezhanbi.index.tolist()*6
indexlist5 = []#top10品牌
for i in range(33,63):
    indexlist5.append('B'+ str(i))
for i in range(67,97):
    indexlist5.append('B'+ str(i))
for i in range(0,len(toplist)):
    ws[indexlist5[i]] = toplist[i]

xiaoelist1 = range(33,43)
xiaoelist2 = list('CDEFGHIJK')
for i in range(10):
    for j in range(9):
        ws[xiaoelist2[j]+str(xiaoelist1[i])] = df_topxiaoezhanbi.iloc[i,j]

junjialist1 = range(67,77)
junjialist2 = list('CDEFGHIJK')
for i in range(10):
    for j in range(9):
        ws[junjialist2[j]+str(junjialist1[i])] = df_topjunjiazhanbi.iloc[i,j]

plusflag = sum([*dictshuxing.values()]) + len(dictshuxing)
startflag = 102
shuxinglist2 = list('CDEFGHIJK')
for z in shuxing_list:
    df_shuxingzhanbi = df.pivot_table(index=z
                                      , columns='周度'
                                      , values='销量'
                                      , aggfunc={'销量': np.sum}
                                      , margins=True).drop(labels='All')
    for i in range(-9, 0):
        df_shuxingzhanbi[df_shuxingzhanbi.columns[i]] = (
                    df_shuxingzhanbi[df_shuxingzhanbi.columns[i]] / df_shuxingzhanbi[df_shuxingzhanbi.columns[i]].sum())
    df_topshuxing = df_shuxingzhanbi.iloc[:10, -9:]

    shuxinglen = dictshuxing[z]
    shuxinglist1 = range(startflag, startflag + shuxinglen)
    for m in range(shuxinglen):
        for n in range(len(shuxinglist2)):
            ws[shuxinglist2[n] + str(shuxinglist1[m])] = df_topshuxing.iloc[m, n]

    shuxinglist3 = [*df_topshuxing.index]
    indexlist6 = []  # 属性
    indexlist7 = []
    indexlist8 = []
    for p in shuxinglist1:
        indexlist6.append('B' + str(p))
        indexlist7.append('B' + str(p + plusflag))
        indexlist8.append('B' + str(p + plusflag + plusflag))
    for q in range(0, len(shuxinglist1)):
        ws[indexlist6[q]] = shuxinglist3[q]
        ws[indexlist7[q]] = shuxinglist3[q]
        ws[indexlist8[q]] = shuxinglist3[q]

    startflag = startflag + shuxinglen + 1

wb.save(path + r"\平台历史数据变更说明.xlsx")
print('success')