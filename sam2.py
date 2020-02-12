import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os

path = os.getcwd()

wb = load_workbook(path + r"\平台历史数据变更说明.xlsx")
ws = wb['平台数据变更说明表']

df2 = pd.read_excel(path + r"\季度变更文档本期.xlsx")

shuxing_list = df2.columns.tolist()[6:]
for i in shuxing_list:
    df2[i] = df2[i].str.upper()
dictshuxing = {}
for i in shuxing_list:
    d = df2[i].drop_duplicates(keep='first', inplace=False)
    dictshuxing[i] = len(d)

bq_weeklist = []
for row in ws['C66':'J66']:
        # 外层for遍历每一行
        for cellObj in row:
            # 内层for遍历改行每一列
            bq_weeklist.append(cellObj.value)
bq_weeklist = bq_weeklist + ['All']

df2_xiaoliang = df2.pivot_table(index = '品类', columns = '周度', values = '销量', aggfunc = np.sum, margins=True)
df2_xiaoliang2 = df2_xiaoliang[bq_weeklist].iloc[:1,:].values.tolist()[0]
for i in range(3,12):
    ws.cell(24,i).value = df2_xiaoliang2[i-3]

df2_xiaoe = df2.pivot_table(index = '品类', columns = '周度', values = '销额', aggfunc = np.sum, margins=True)
df2_xiaoe2 = df2_xiaoe[bq_weeklist].iloc[:1,:].values.tolist()[0]
df2_xiaoe3 = np.array(df2_xiaoe2)/10000
for i in range(3,12):
    ws.cell(25,i).value = df2_xiaoe3[i-3]

df2_junjia = df2_xiaoe/df2_xiaoliang
df2_junjia2 = df2_junjia[bq_weeklist].iloc[:1,:].values.tolist()[0]
for i in range(3,12):
    ws.cell(26,i).value = df2_junjia2[i-3]

df2_xiaoejunjiazhanbi = df2.pivot_table(index = '品牌'
                                      , columns = '周度'
                                      , values = ['销额','销量']
                                      , aggfunc = {'销额':np.sum,'销量':np.sum}
                                      , margins=True).sort_values(('销额',bq_weeklist[-2]),ascending=False).drop(labels = 'All')

df2_xiaoezhanbi = df2_xiaoejunjiazhanbi['销额']
df2_junjiazhanbi = df2_xiaoejunjiazhanbi['销额']/df2_xiaoejunjiazhanbi['销量']

df2_xiaoezhanbi2 = df2_xiaoezhanbi[bq_weeklist]
df2_junjiazhanb2 = df2_junjiazhanbi[bq_weeklist]

for i in range(-9,0):
    df2_xiaoezhanbi2[df2_xiaoezhanbi2.columns[i]] = (df2_xiaoezhanbi2[df2_xiaoezhanbi2.columns[i]] / df2_xiaoezhanbi2[df2_xiaoezhanbi2.columns[i]].sum())

bq_pinpailist = []
for row in ws['B33':'B42']:
    # 外层for遍历每一行
    for cellObj in row:
        # 内层for遍历改行每一列
        bq_pinpailist.append(cellObj.value)

topxiaoelist = []
for i in bq_pinpailist:
    topxiaoelist.append(df2_xiaoezhanbi2.loc[i].values.tolist())

for i in range(43,53):
    for j in range(3,12):
        ws.cell(i,j).value = topxiaoelist[i-43][j-3]#销额

topjunjialist = []
for i in bq_pinpailist:
    topjunjialist.append(df2_junjiazhanb2.loc[i].values.tolist())

for i in range(77,87):
    for j in range(3,12):
        ws.cell(i,j).value = topjunjialist[i-77][j-3]#均价

plusflag = sum([*dictshuxing.values()])+len(dictshuxing)
startflag = 101 + plusflag + 1

for z in shuxing_list:
    df2_shuxing1_1 = df2.pivot_table(index=z, columns='周度', values='销量', aggfunc=np.sum, margins=True).sort_values(
        bq_weeklist[-2], ascending=False).drop(labels='All')
    df2_shuxing1_2 = df2_shuxing1_1[bq_weeklist]
    for i in range(-9, 0):
        df2_shuxing1_2[df2_shuxing1_2.columns[i]] = (
                    df2_shuxing1_2[df2_shuxing1_2.columns[i]] / df2_shuxing1_2[df2_shuxing1_2.columns[i]].sum())

    x1 = startflag
    y1 = startflag + dictshuxing[z]
    x2 = 'B' + str(startflag)
    y2 = 'B' + str(startflag + dictshuxing[z] - 1)

    startflag = startflag + dictshuxing[z] + 1

    bq_shuxinglist = []
    for row in ws[x2:y2]:
        # 外层for遍历每一行
        for cellObj in row:
            # 内层for遍历改行每一列
            bq_shuxinglist.append(cellObj.value)

    topshuxinglist = []
    for i in bq_shuxinglist:
        topshuxinglist.append(df2_shuxing1_2.loc[i].values.tolist())
    for i in range(x1, y1):
        for j in range(3, 12):
            ws.cell(i, j).value = topshuxinglist[i - x1][j - 3]  # 属性1

wb.save(path + r"\平台历史数据变更说明.xlsx")
print('success')